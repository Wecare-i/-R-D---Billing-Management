"""
Fill Expense Report Script
--------------------------
Đọc data từ CSV đã fill → ghi vào file Excel mới.
Giữ nguyên template gốc, tạo file output mới.

Usage:
  python scripts/fill_expense.py <csv_file> [--month MM] [--year YYYY]

Ví dụ:
  python scripts/fill_expense.py "Bảng kê chi phí_filled.csv" --month 03 --year 2026
  → Output: output/Tech_Bảng kê chi phí_T03.2026.xlsx
"""

import csv
import sys
import os
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, numbers
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def read_csv_data(csv_path):
    """Đọc CSV và trả về list of rows."""
    rows = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    return rows


def csv_to_excel(csv_path, output_path, template_xlsx=None):
    """
    Convert CSV data sang Excel.
    Nếu có template_xlsx: copy template rồi ghi data vào.
    Nếu không có template: tạo Excel mới từ CSV data.
    """
    if not HAS_OPENPYXL:
        print("ERROR: Cần cài openpyxl. Chạy: pip install openpyxl")
        sys.exit(1)

    rows = read_csv_data(csv_path)

    if template_xlsx and os.path.exists(template_xlsx):
        # Copy template và ghi data vào
        print(f"📋 Dùng template: {template_xlsx}")
        wb = openpyxl.load_workbook(template_xlsx)
        ws = wb.active

        # Ghi data từ CSV vào đúng vị trí trong template
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                # Chỉ ghi nếu có data (không ghi đè format của ô trống)
                if value.strip():
                    clean_val = value.replace(',', '')
                    try:
                        # Thử parse float trước để giữ decimal (VD: $1,160.87 → 1160.87)
                        float_val = float(clean_val)
                        if float_val == int(float_val) and '.' not in value:
                            cell.value = int(float_val)
                            cell.number_format = '#,##0'
                        else:
                            cell.value = float_val
                            cell.number_format = '#,##0.00'
                    except ValueError:
                        cell.value = value
    else:
        # Tạo Excel mới từ CSV
        print("📝 Tạo Excel mới từ CSV (không có template)")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bảng kê chi phí"

        # Setup styles
        header_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if value.strip():
                    clean_val = value.replace(',', '')
                    try:
                        # Thử parse float trước để giữ decimal (VD: $1,160.87 → 1160.87)
                        float_val = float(clean_val)
                        if float_val == int(float_val) and '.' not in value:
                            cell.value = int(float_val)
                            cell.number_format = '#,##0'
                        else:
                            cell.value = float_val
                            cell.number_format = '#,##0.00'
                    except ValueError:
                        cell.value = value

    # Tạo thư mục output nếu chưa có
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb.save(output_path)
    print(f"✅ Đã tạo file: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Fill expense report CSV → Excel')
    parser.add_argument('csv_file', help='Path to filled CSV file')
    parser.add_argument('--month', '-m', help='Tháng (MM)', default=datetime.now().strftime('%m'))
    parser.add_argument('--year', '-y', help='Năm (YYYY)', default=datetime.now().strftime('%Y'))
    parser.add_argument('--template', '-t', help='Path to Excel template (.xlsx)', default=None)

    args = parser.parse_args()

    # Build output filename
    output_name = f"Tech_Bảng kê chi phí_T{args.month}.{args.year}.xlsx"
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    output_path = os.path.join(project_dir, 'output', output_name)

    # Check template
    template_path = args.template
    if not template_path:
        # Auto-detect template in templates/
        default_template = os.path.join(project_dir, 'templates', 'Bảng kê chi phí.xlsx')
        if os.path.exists(default_template):
            template_path = default_template

    # Check output already exists
    if os.path.exists(output_path):
        response = input(f"⚠️ File {output_name} đã tồn tại. Ghi đè? (y/n): ")
        if response.lower() != 'y':
            print("❌ Hủy.")
            return

    csv_to_excel(args.csv_file, output_path, template_path)


if __name__ == '__main__':
    main()
